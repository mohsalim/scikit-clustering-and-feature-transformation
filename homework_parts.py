from bench_mark import bench_kmeans, bench_em, bench_pca, bench_ica, bench_rca, bench_lda, bench_etr
from openpyxl import Workbook
from sklearn.cluster import KMeans
from sklearn.random_projection import GaussianRandomProjection
from sklearn import mixture
from sklearn import metrics
from sklearn.decomposition import PCA
from sklearn.decomposition import FastICA
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
import time
import numpy as np
from sklearn.neural_network import MLPRegressor
from sklearn.tree import ExtraTreeRegressor
from sklearn.feature_selection import SelectFromModel

def part1(data, target, x_train, y_train, x_test, y_test):
    # K-Means
    # TODO find best n_clusters range (1-50?)
    print('--- KMeans ---')
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'train wall time', 'train accuracy', 'test wall time', 'test accuracy', 'inertia', 'homogeneity', 'completeness', 'v measure', 'ARI', 'AMI', 'FMI']
    ws.append(headers)
    for n in range(1, 31):
        bench_mark = bench_kmeans(KMeans(n_clusters=n, random_state=0), 'KMeans', n, x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-1-kmeans-bench.xlsx')

    # EM
    # Run each covariance_type (defaults to full).
    # full: each component has its own general covariance matrix.
    # tied: all components share the same general covariance matrix.
    # diag: each component has its own diagonal covariance matrix.
    # spherical: each component has its own single variance.
    print('--- EM ---')
    cv_types = ['spherical', 'tied', 'diag', 'full']
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'train wall time', 'train score', 'test wall time', 'test accuracy', 'AIC', 'BIC']
    ws.append(headers)
    for cv_type in cv_types:
        # For each n size components.
        for n in range(1, 31):
            # Run expectation maximization (EM) algorithm.
            # Note: Gaussian Mixture implements EM.
            bench_mark = bench_em(mixture.GaussianMixture(n_components=n, covariance_type=cv_type), 'EM ' + cv_type, n, x_train, y_train, x_test, y_test)
            ws.append(bench_mark)
    wb.save('part-1-em-bench.xlsx')

def part2(data, target, x_train, y_train, x_test, y_test, features_count):
    fc = features_count + 1
    
    # PCA
    print('--- PCA ---')
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'train wall time', 'noise variance', 'dimension size after dimensionality reduction',
               'components', 'explained variance', 'explained variance ratio', 'mean', 'n components']
    ws.append(headers)
    for n in range(1, fc):
        bench_mark = bench_pca(PCA(n_components=n), 'PCA', n, x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-2-pca-bench.xlsx')

    # ICA
    print('--- ICA ---')
    # TODO consider trying different functions (fun) and tolerance values (tol).
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'train wall time', 'iterations', 'dimension size after dimensionality reduction',
               'components (unmixing matrix)', 'mixing matrix',
               'original kurtosis', 'ok min', 'ok max', 'ok average',
               'transformed kurtosis', 'tk min', 'tk max', 'tk average',
               'components kurtosis', 'ck min', 'ck max', 'ck average',
               'mixing kurtosis', 'mk min', 'mk max', 'mk average']
    ws.append(headers)
    for n in range(1, fc):
        bench_mark = bench_ica(FastICA(n_components=n), 'ICA', n, x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-2-ica-bench.xlsx')

    # RCA
    print('--- RCA ---')
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'iterations', 'train wall time', 'components', 'n components']
    ws.append(headers)
    for n in range(1, fc):
        for i in range(0, 101):
            bench_mark = bench_rca(GaussianRandomProjection(n_components=n), 'RCA', n, i, x_train, y_train, x_test, y_test)
            ws.append(bench_mark)
    wb.save('part-2-rca-bench.xlsx')

    # LDA
    print('--- LDA ---')
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'k', 'train wall time', 'coef', 'intercept',
               'explained variance ratio', 'means', 'priors', 'scalings', 'xbar', 'classes']
    ws.append(headers)
    for n in range(1, fc):
        bench_mark = bench_lda(LinearDiscriminantAnalysis(n_components=n), 'LDA', n, x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-2-lda-bench.xlsx')

    # ETR (Extra Trees Regressor)
    print('--- ETR ---')
    wb = Workbook()
    ws = wb.active
    headers = ['algorithm', 'train wall time', 'old feature size', 'new feature size', 'feature importances']
    ws.append(headers)
    bench_mark = bench_etr(ExtraTreeRegressor(), 'ETR', x_train, y_train, x_test, y_test)
    ws.append(bench_mark)
    wb.save('part-2-etr-bench.xlsx')

def prepend_headers(algo, headers):
    new_headers = []
    for header in headers:
        new_headers.append(algo + ' ' + header)
    return new_headers

def part3(data, target, k_list, x_train, y_train, x_test, y_test):
    headers = ['algorithm', 'k', 'train wall time', 'test wall time', 'training accuracy', 'testing accuracy', 'training MSE', 'testing MSE']
    
    # PCA -> KMeans
    print('--- PCA -> KMeans ---')
    wb = Workbook()
    ws = wb.active
    ws.append(prepend_headers('KMeans PCA', headers))
    for k in k_list:
        pca = PCA(n_components=k).fit(x_train)
        # In this case the seeding of the centers is deterministic.
        # Hence we run the k-means algorithm only once with n_init=1.
        # Else wise it will give a warning and set n_init to 1 anyways.
        kmeans = KMeans(init=pca.components_, n_clusters=k, n_init=1, random_state=0)
        bench_mark = run_kmeans_with_dr(kmeans, k, 'KMeans', 'PCA', x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-3-pca-kmeans-bench.xlsx')

    # ICA -> KMeans
    print('--- ICA -> KMeans ---')
    wb = Workbook()
    ws = wb.active
    ws.append(prepend_headers('KMeans ICA', headers))
    for k in k_list:
        ica = FastICA(n_components=k).fit(x_train)
        kmeans = KMeans(init=ica.components_, n_clusters=k, n_init=1, random_state=0)
        bench_mark = run_kmeans_with_dr(kmeans, k, 'KMeans', 'ICA', x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-3-ica-kmeans-bench.xlsx')

    # RCA -> KMeans
    print('--- RCA -> KMeans ---')
    wb = Workbook()
    ws = wb.active
    ws.append(prepend_headers('KMeans RCA', headers))
    for k in k_list:
        rca = GaussianRandomProjection(n_components=k).fit(x_train)
        kmeans = KMeans(init=rca.components_, n_clusters=k, n_init=1, random_state=0)
        bench_mark = run_kmeans_with_dr(kmeans, k, 'KMeans', 'RCA', x_train, y_train, x_test, y_test)
        ws.append(bench_mark)
    wb.save('part-3-rca-kmeans-bench.xlsx')

    # ETR -> KMeans
    print('--- ETR -> KMeans ---')
    wb = Workbook()
    ws = wb.active
    ws.append(prepend_headers('KMeans ETR', headers))
    x_train_float = x_train.astype(np.float)
    y_train_int = y_train.astype(int)
    for k in k_list:
        etr = ExtraTreeRegressor().fit(x_train_float, y_train_int)
        model = SelectFromModel(etr, prefit=True)
        x_train_new = model.transform(x_train_float)
        x_test_new = model.transform(x_test)
        kmeans = KMeans(n_clusters=k, n_init=1, random_state=0)
        bench_mark = run_kmeans_with_dr(kmeans, k, 'KMeans', 'ETR', x_train_new, y_train, x_test_new, y_test)
        ws.append(bench_mark)
    wb.save('part-3-etr-kmeans-bench.xlsx')

    cv_types = ['spherical', 'tied', 'diag', 'full']
    for cv_type in cv_types:
        algo_type = 'EM ' + cv_type
        algo_type_file = 'em-' + cv_type

        # PCA -> EM
        print('--- PCA -> EM ---')
        wb = Workbook()
        ws = wb.active
        ws.append(prepend_headers(algo_type + ' PCA', headers))
        for k in k_list:
            pca = PCA(n_components=k).fit(x_train)
            # TODO should I be using weights_init or means_init
            em = mixture.GaussianMixture(n_components=k, covariance_type=cv_type, means_init=pca.components_)
            bench_mark = run_em_with_dr(em, k, algo_type, 'PCA', x_train, y_train, x_test, y_test)
            ws.append(bench_mark)
        wb.save('part-3-pca-' + algo_type_file + '-bench.xlsx')

        # ICA -> EM
        print('--- ICA -> EM ---')
        wb = Workbook()
        ws = wb.active
        ws.append(prepend_headers(algo_type + ' ICA', headers))
        for k in k_list:
            ica = FastICA(n_components=k).fit(x_train)
            em = mixture.GaussianMixture(n_components=k, covariance_type=cv_type, means_init=ica.components_)
            bench_mark = run_em_with_dr(em, k, algo_type, 'ICA', x_train, y_train, x_test, y_test)
            ws.append(bench_mark)
        wb.save('part-3-ica-' + algo_type_file + '-bench.xlsx')

        # RCA -> EM
        print('--- RCA -> EM ---')
        wb = Workbook()
        ws = wb.active
        ws.append(prepend_headers(algo_type + ' RCA', headers))
        for k in k_list:
            rca = GaussianRandomProjection(n_components=k).fit(x_train)
            em = mixture.GaussianMixture(n_components=k, covariance_type=cv_type, means_init=rca.components_)
            bench_mark = run_em_with_dr(em, k, algo_type, 'RCA', x_train, y_train, x_test, y_test)
            ws.append(bench_mark)
        wb.save('part-3-rca-' + algo_type_file + '-bench.xlsx')

        # ETR -> EM
        print('--- ETR -> EM ---')
        wb = Workbook()
        ws = wb.active
        ws.append(prepend_headers(algo_type + ' ETR', headers))
        x_train_float = x_train.astype(np.float)
        y_train_int = y_train.astype(int)
        for k in k_list:
            etr = ExtraTreeRegressor().fit(x_train_float, y_train_int)
            model = SelectFromModel(etr, prefit=True)
            x_train_new = model.transform(x_train_float)
            x_test_new = model.transform(x_test)
            em = mixture.GaussianMixture(n_components=k, covariance_type=cv_type)
            bench_mark = run_em_with_dr(kmeans, k, algo_type, 'ETR', x_train_new, y_train, x_test_new, y_test)
            ws.append(bench_mark)
        wb.save('part-3-etr-' + algo_type_file + '-bench.xlsx')

def run_kmeans_with_dr(cluster_algo, k, clustering_name, dr_name, x_train, y_train, x_test, y_test):
    # Train.
    start = time.time()
    cluster_algo.fit(x_train)
    train_time = time.time() - start
    train_score = metrics.accuracy_score(y_train, cluster_algo.labels_)
    train_mse = metrics.mean_squared_error(y_train, cluster_algo.labels_)
    algo_name = clustering_name + " " + dr_name
    #print(algo_name + " " + str(k) + " train: " + str(score))
    
    #Test
    start = time.time()
    predicted_labels = cluster_algo.predict(x_test)
    test_time = time.time() - start
    test_score = metrics.accuracy_score(y_test, predicted_labels)
    test_mse = metrics.mean_squared_error(y_test, predicted_labels)
    #print(algo_name + " " + str(k) + " test: " + str(score))

    return (algo_name, k, train_time, test_time, train_score, test_score, train_mse, test_mse)

def run_em_with_dr(cluster_algo, k, clustering_name, dr_name, x_train, y_train, x_test, y_test):
    # Train.
    algo_name = clustering_name + " " + dr_name
    start = time.time()
    cluster_algo.fit(x_train)
    train_time = time.time() - start
    predicted_labels = cluster_algo.predict(x_train)
    train_score = metrics.accuracy_score(y_train, predicted_labels)
    train_mse = metrics.mean_squared_error(y_train, predicted_labels)
    #print(algo_name + " " + str(k) + " test: " + str(score))

    #Test
    start = time.time()
    predicted_labels = cluster_algo.predict(x_test)
    test_time = time.time() - start
    test_score = metrics.accuracy_score(y_test, predicted_labels)
    test_mse = metrics.mean_squared_error(y_test, predicted_labels)
    #print(algo_name + " " + str(k) + " test: " + str(score))
    
    return (algo_name, k, train_time, test_time, train_score, test_score, train_mse, test_mse)

def part4(data, target, k_list, x_train, y_train, x_test, y_test):
    x_train_float = x_train.astype(np.float)
    y_train_int = y_train.astype(int)
    x_test_float = x_test.astype(np.float)
    
    # PCA -> ANN
    print('--- PCA -> ANN ---')
    for k in k_list:
        pca = PCA(n_components=k).fit(x_train)
        mlp = MLPRegressor(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(15,), random_state=1)
        run_ann_with_dr(mlp, k, 'ANN', 'PCA', pca.transform(x_train_float), y_train, pca.transform(x_test_float), y_test)

    # ICA -> ANN
    print('--- ICA -> ANN ---')
    for k in k_list:
        ica = FastICA(n_components=k).fit(x_train)
        mlp = MLPRegressor(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(15,), random_state=1)
        run_ann_with_dr(mlp, k, 'ANN', 'ICA', ica.transform(x_train_float), y_train, ica.transform(x_test_float), y_test)

    # RCA -> ANN
    print('--- RCA -> ANN ---')
    for k in k_list:
        rca = GaussianRandomProjection(n_components=k).fit(x_train)
        mlp = MLPRegressor(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(15,), random_state=1)
        run_ann_with_dr(mlp, k, 'ANN', 'RCA', rca.transform(x_train_float), y_train, rca.transform(x_test_float), y_test)

    # ETR -> ANN
    print('--- ETR -> ANN ---')
    for k in k_list:
        etr = ExtraTreeRegressor().fit(x_train_float, y_train_int)
        model = SelectFromModel(etr, prefit=True)
        mlp = MLPRegressor(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(15,), random_state=1)
        run_ann_with_dr(mlp, k, 'ANN', 'ETR', model.transform(x_train_float), y_train, model.transform(x_test_float), y_test)

def run_ann_with_dr(ann, k, ann_name, dr_name, x_train, y_train, x_test, y_test):
    # Train.
    ann.fit(x_train, y_train)
    score = metrics.mean_squared_error(y_train.astype(np.float), ann.predict(x_train))
    algo_name = ann_name + " " + dr_name
    print(algo_name + " " + str(k) + " train: " + str(score))
    
    # Test.
    score = metrics.mean_squared_error(y_test.astype(np.float), ann.predict(x_test)) 
    print(algo_name + " " + str(k) + " test: " + str(score))


def part5(data, target, k_list, x_train, y_train, x_test, y_test):
    x_train_float = x_train.astype(np.float)
    y_train_int = y_train.astype(int)
    x_test_float = x_test.astype(np.float)
    
    # After analyzing the data from part 3, these are probably the best Cluster + DR algorithms to run:

    # KMeans + ETR, k = 2
    print('--- ETR -> KMeans -> ANN ---')
    k = 2
    etr = ExtraTreeRegressor().fit(x_train_float, y_train_int)
    model = SelectFromModel(etr, prefit=True)
    x_train_new = model.transform(x_train_float)
    x_test_new = model.transform(x_test_float)
    kmeans = KMeans(n_clusters=k, n_init=1, random_state=0)
    train_cluster = kmeans.fit_transform(x_train_new.astype(np.float), y_train.astype(int))
    test_cluster = kmeans.fit_transform(x_test_new.astype(np.float), y_test.astype(int))
    mlp = MLPRegressor(solver='lbfgs', alpha=1e-5, hidden_layer_sizes=(15,), random_state=1)
    run_ann_with_dr(mlp, k, 'ANN', 'PCA', train_cluster, y_train, test_cluster, y_test)

    # EM Diag + ICA, k = 2

    # EM Full + ICA, k = 2

    # KMeans + PCA, k = 2

    # EM Tied + PCA, k = 3

    # EM Tied + RCA, k = 2

    # EM Diag + RCA, k = 3

    # EM Tied + RCA, k = 27
